from django.template.response import TemplateResponse
from django.http import HttpResponseRedirect, HttpResponse

from aplicacoes.usuario.views import verificacao_maxima
from aplicacoes.atena.atualizar_turmas import *
from .models import *
from aplicacoes.administracao.models import *
from aplicacoes.lotacao.models import *
import pandas as pd
import datetime
from aplicacoes.lotacao.models import *
from openpyxl import *
import os
from django.conf import settings

from django.template.loader import get_template
from xhtml2pdf import pisa
from aplicacoes.core.models import Confirmacao_lotacao
from aplicacoes.atena.importar_simaed import *


def index(request):
    if not verificacao_maxima(request, [7]):
        return HttpResponseRedirect('/')

    user = request.session['username']
    template_name = 'atena/index.html'

    detalhe = Detalhes.objects.get(id= 1)

    if user == 'joaoteixeira.nettoasd':
        lista = ['AB.xlsx', 'ACRELANDIA.xlsx', 'BRASILEIA.xlsx', 'BUJARI.xlsx', 'CAPIXABA.xlsx', 'CZS.xlsx', 'EPITACIOLANDIA.xlsx', 'FEIJO.xlsx', 'JORDAO.xlsx', 'MB.xls', 'MT.xls', 'PA.xls', 'PDC.xlsx', 'PW.xls', 'RA.xlsx', 'SA.xls', 'SG.xlsx', 'SM.xlsx', 'TARAUCA.xlsx', 'XAPURI.xlsx']

        matriculas = []
        digitos = []
        subs = []
        amparos = []
        f1 = []
        f2 = []

        for planilha in lista:
            dado = pd.read_excel(f'aplicacoes/atena/lotacao/{planilha}')
            v1 = dado['MATRICULA'].tolist()
            v2 = dado['DG'].tolist()
            v3 = dado['SUB'].tolist()
            v4 = dado['AMPARO LEGAL'].tolist()
            v5 = dado['FORMACAO'].tolist()
            v6 = dado['2 FORMACAO'].tolist()

            for i in range(len(v1)):
                matriculas.append(v1[i])

                try:
                    digitos.append(str(int(v2[i])))
                except:
                    digitos.append(v2[i])

                try:
                    subs.append(str(int(v3[i])))
                except:
                    subs.append(v3[i])

                amparos.append(v4[i])
                f1.append(v5[i])
                f2.append(v6[i])

        for item in range(len(matriculas)):
            servidor = Servidor.objects.filter(matricula= matriculas[item])
            if servidor.count() == 1:
                servidor = servidor[0]
                if Servidor_contrato.objects.filter(servidor= servidor, digito= digitos[item]).exists():
                    contrato = Servidor_contrato.objects.get(servidor= servidor, digito= digitos[item])
                    amparo = amparos[item]

                    try:
                        if 'DOE' in amparo:
                            doe = amparo.replace('DOE ', '')
                            contrato.doe = doe
                            contrato.save()
                        elif 'PARECER' in amparo:
                            parecer = amparo.replace('PARECER ', '')
                            contrato.parecer = parecer
                            contrato.save()
                    except:
                        pass

                    for lotacao in Servidor_lotacao.objects.filter(contrato= contrato):
                        if Servidor_subconta.objects.filter(sub= subs[item]).exists():
                            lotacao.subconta = Servidor_subconta.objects.get(sub= subs[item])
                            lotacao.save()

                    if not Servidor_contrato_formacao.objects.filter(contrato= contrato, tipo= '1', formacao= f1[item]).exists():
                        Servidor_contrato_formacao(contrato= contrato, tipo= '1', formacao= f1[item]).save()

                    if not Servidor_contrato_formacao.objects.filter(contrato= contrato, tipo= '2', formacao= f1[item]).exists():
                        Servidor_contrato_formacao(contrato= contrato, tipo= '2', formacao= f1[item]).save()

    if request.method == 'POST':
        if request.POST.get('botao'):
            valor_botao = request.POST.get('botao')
            detalhe.situacao = valor_botao
            detalhe.save()
        else:
            detalhe.atualizar_simaed = 1
            detalhe.situacao = 'Desativado'
            detalhe.save()

            handle_uploaded_file(request.FILES.get('arquivo'), 'dados_simaed.csv', 'Atena/', 'Atualizar dados do SIMAED')
            ler_planilha()

            detalhe.atualizar_simaed = 0
            detalhe.situacao = 'Ativo'
            detalhe.save()
            return HttpResponseRedirect('/atena')

    if user == 'joaopsedro.passos':



        dados = pd.read_excel(r"C:/Users/joao.soares/Documents/code/folha.xlsx", )


        ord_coluna = dados["ORD"].to_list()
        matricula = dados["MATRICULA"].to_list()
        dg = dados["DG"].to_list()
        nome = dados["NOME DO SERVIDOR"].to_list()
        cargos = dados["CARGO"].to_list()
        adimissão = dados["ADMISSÃO"].to_list()
        contrato = dados["CONTRATO"].to_list()
        cpf = dados["C.P.F"].to_list()
        sub = dados["SUB"].to_list()
        pis = dados["PIS/PASEP"].to_list()
        nascimento = dados["NASCIMENTO"].to_list()
        ref = dados["REF"].to_list()
        sexo = dados["SEXO"].to_list()
        lotacao = dados["LOTAÇÃO"].to_list()
        municipio = dados["MUNICIPIO"].to_list()
        situacao = dados["SITUAÇÃO DO SERVIDOR"].to_list()
        vencimento_base = dados["VENCIMENTO BASE"].to_list()
        vencimento_bruto = dados["VENCIMENTO BRUTO"].to_list()

        valores = []

        for item in range(len(cargos)):
            if (f"{cargos[item]}-{ref[item]}-{vencimento_base[item]}") not in valores and ref[item] != 0:
                valores.append(f"{cargos[item]}-{ref[item]}-{vencimento_base[item]}")

                if Cargo.objects.filter(nome = cargos[item]).exists():
                    cargo_vencimento = Cargo_vencimento()
                    cargos_servidores = Cargo.objects.get(nome = cargos[item])

                    cargo_vencimento.ref = ref[item]
                    cargo_vencimento.cargo = cargos_servidores
                    cargo_vencimento.valor = vencimento_base[item]
                    # cargo_vencimento.save()
                    print(cargos_servidores,' ',cargos_servidores.id, ' ', ref[item], ' ', vencimento_base[item])

    # pasta_raiz = os.getcwd()

    # pasta_css = os.path.join(pasta_raiz, 'static', 'assets', 'css')
    # pasta_templates = os.path.join(pasta_raiz, 'templates')

    # print(pasta_css)
    # print(pasta_templates)

    # print(os.listdir(pasta_css))

    return TemplateResponse(request, template_name, locals())

def teste_escolas_etapas(request):
    escolas_etapas = Etapa_escola.objects.filter(etapa_id= 10).values_list('escola__id', flat= True)
    escolas = Escola.objects.exclude(id__in= escolas_etapas)
    etapa_aee = Etapa.objects.get(id= 10)
    len = escolas.count()

    # for escola in escolas:
    #     novo = Etapa_escola()
    #     novo.escola = escola
    #     novo.etapa = etapa_aee
    #     novo.save()



    # print()

    template_name = 'atena/teste-escolas-etapas.html'

    return TemplateResponse(request, template_name, locals())

def teste(request):
    if not verificacao_maxima(request, [7]):
        return HttpResponseRedirect('/')

    user = request.session['username']

    pagina = request.GET.get('pagina')

    template_name = f'atena/{pagina}.html'

    return TemplateResponse(request, template_name, locals())

def aditivo(request):
    aditivos = Servidor_contrato_aditivo.objects.all()

    for aditivo in aditivos:
        data_termino = str(aditivo.contrato.data_termino)

        if data_termino != 'None':
            print(data_termino)
            dia = data_termino.split('-')[2]
            mes = data_termino.split('-')[1]
            ano = data_termino.split('-')[0]

            dia_int = int(dia)
            mes_int = int(mes)
            ano_int = int(ano)



            if (dia_int < 28 and mes_int == 2) or (dia_int < 30 and mes_int != 2):
                print(data_termino, ano_int, mes_int, dia_int)
                data_nova = datetime.datetime(ano_int, mes_int, dia_int+1)

                print(data_nova)
                aditivo.data_inicio = data_nova

                print(aditivo.data_inicio)
                aditivo.save()


        print()


    return HttpResponseRedirect('/atena')

def front(request):
    if not verificacao_maxima(request, [7]):
        return HttpResponseRedirect('/')

    user = request.session['username']
    template_name = 'atena/front.html'

    return TemplateResponse(request, template_name, locals())

def testes(request):
    if not verificacao_maxima(request, [7]):
        return HttpResponseRedirect('/')

    user = request.session['username']
    template_name = 'atena/testes.html'

    # AQUI VALE TUDO MEU AMIGO, CADA UM ESCREVE SEU TESTE E BOA SORTE

    # FRANKLEEEEEEEEEEEEE
    if user == 'franklin.fariasss':
        urls = str(settings.BASE_DIR)

        # Abre o arquivo pelo caminho do documento.
        wb = load_workbook(filename= urls + '/aplicacoes/dinem/relatorios/relatorio_aluno_parcial_2022.xlsx')

        # Identifaca a aba da planilha pelo nome.
        sh = wb['Planilha1']

        # Dados que serão inseridos na planilha.
        nome = 'Franklin Farias'
        nascimento = '13/09/2000'
        naturalidade = 'Rio Branco'
        uf = 'AC'
        pai = 'Francisco Evandro'
        mae = 'Maria Elenir'

        sh.cell(row= 3, column= 2).value = (sh.cell(row= 3, column= 2).value).replace('aluno.nome', nome)
        sh.cell(row= 3, column= 2).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 4, column= 1).value = (sh.cell(row= 4, column= 1).value).replace('aluno.data_nascimento', nascimento)
        sh.cell(row= 4, column= 1).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 4, column= 4).value = (sh.cell(row= 4, column= 4).value).replace('aluno.naturalidade', naturalidade)
        sh.cell(row= 4, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 4, column= 8).value = (sh.cell(row= 4, column= 8).value).replace('aluno.uf', uf)
        sh.cell(row= 4, column= 8).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 5, column= 2).value = (sh.cell(row= 5, column= 2).value).replace('aluno.pai', pai)
        sh.cell(row= 5, column= 2).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 6, column= 2).value = (sh.cell(row= 6, column= 2).value).replace('aluno.mae', mae)
        sh.cell(row= 6, column= 2).font = styles.Font(name= 'Calibri', size= 9)


        portugues = '9.7'
        artes = '10'
        ingles = '7.8'
        ed_fisica= '10'
        matematica = '8.5'
        fisica = '8.6'
        quimica = '8.7'
        biologia = '8.8'
        historia = '8.9'
        geografia = '8.1'
        filosofia = '8.2'
        sociologia = '8.3'


        #Relatorio1
        sh.cell(row= 9, column= 4).value = (sh.cell(row= 9, column= 4).value).replace('relatorio1.portugues', portugues)
        sh.cell(row= 9, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 10, column= 4).value = (sh.cell(row= 10, column= 4).value).replace('relatorio1.artes', artes)
        sh.cell(row= 10, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 11, column= 4).value = (sh.cell(row= 11, column= 4).value).replace('relatorio1.ingles', ingles)
        sh.cell(row= 11, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 12, column= 4).value = (sh.cell(row= 12, column= 4).value).replace('relatorio1.ed_fisica', ed_fisica)
        sh.cell(row= 12, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 13, column= 4).value = (sh.cell(row= 13, column= 4).value).replace('relatorio1.matematica', matematica)
        sh.cell(row= 13, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 14, column= 4).value = (sh.cell(row= 14, column= 4).value).replace('relatorio1.fisica', fisica)
        sh.cell(row= 14, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 15, column= 4).value = (sh.cell(row= 15, column= 4).value).replace('relatorio1.quimica', quimica)
        sh.cell(row= 15, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 16, column= 4).value = (sh.cell(row= 16, column= 4).value).replace('relatorio1.biologia', biologia)
        sh.cell(row= 16, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 17, column= 4).value = (sh.cell(row= 17, column= 4).value).replace('relatorio1.historia', historia)
        sh.cell(row= 17, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 18, column= 4).value = (sh.cell(row= 18, column= 4).value).replace('relatorio1.geografia', geografia)
        sh.cell(row= 18, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 19, column= 4).value = (sh.cell(row= 19, column= 4).value).replace('relatorio1.filosofia', filosofia)
        sh.cell(row= 19, column= 4).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 20, column= 4).value = (sh.cell(row= 20, column= 4).value).replace('relatorio1.sociologia', sociologia)
        sh.cell(row= 20, column= 4).font = styles.Font(name= 'Calibri', size= 9)



        #Relatorio2
        sh.cell(row= 9, column= 6).value = (sh.cell(row= 9, column= 6).value).replace('relatorio2.portugues', portugues)
        sh.cell(row= 9, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 10, column= 6).value = (sh.cell(row= 10, column= 6).value).replace('relatorio2.artes', artes)
        sh.cell(row= 10, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 11, column= 6).value = (sh.cell(row= 11, column= 6).value).replace('relatorio2.ingles', ingles)
        sh.cell(row= 11, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 12, column= 6).value = (sh.cell(row= 12, column= 6).value).replace('relatorio2.ed_fisica', ed_fisica)
        sh.cell(row= 12, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 13, column= 6).value = (sh.cell(row= 13, column= 6).value).replace('relatorio2.matematica', matematica)
        sh.cell(row= 13, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 14, column= 6).value = (sh.cell(row= 14, column= 6).value).replace('relatorio2.fisica', fisica)
        sh.cell(row= 14, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 15, column= 6).value = (sh.cell(row= 15, column= 6).value).replace('relatorio2.quimica', quimica)
        sh.cell(row= 15, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 16, column= 6).value = (sh.cell(row= 16, column= 6).value).replace('relatorio2.biologia', biologia)
        sh.cell(row= 16, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 17, column= 6).value = (sh.cell(row= 17, column= 6).value).replace('relatorio2.historia', historia)
        sh.cell(row= 17, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 18, column= 6).value = (sh.cell(row= 18, column= 6).value).replace('relatorio2.geografia', geografia)
        sh.cell(row= 18, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 19, column= 6).value = (sh.cell(row= 19, column= 6).value).replace('relatorio2.filosofia', filosofia)
        sh.cell(row= 19, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 20, column= 6).value = (sh.cell(row= 20, column= 6).value).replace('relatorio2.sociologia', sociologia)
        sh.cell(row= 20, column= 6).font = styles.Font(name= 'Calibri', size= 9)

        # Relatorio3
        sh.cell(row= 9, column= 8).value = (sh.cell(row= 9, column= 8).value).replace('relatorio3.portugues', portugues)
        sh.cell(row= 9, column= 14).font = styles.Font(name= 'Calibri', size= 9)

        sh.cell(row= 13, column= 8).value = (sh.cell(row= 13, column= 8).value).replace('relatorio3.matematica', matematica)
        sh.cell(row= 13, column= 8).font = styles.Font(name= 'Calibri', size= 9)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Relatório Aluno.xlsx"'

        # Salva os dados na planilha.
        wb.save(response)

        return response


    if user == 'franklin.farias' or user == 'erick.nascimento':
        # importar_turmas()
        # importar_alunos()
        # enturmar_alunos()

        # PASSANDO TODOS OS DADOS DA PLANILHA DO SIMAED PAR TABELA AUX_SIMAED
        # def foritem(x, lista):
        #     for i in x:
        #         lista.append(str(i).replace('nan', ''))

        # def usuarios():
        #     municipio = []
        #     escola = []
        #     inep_escola = []
        #     nome_aluno = []
        #     inep_aluno = []
        #     nascimento = []
        #     nome_pai = []
        #     nome_mae = []
        #     sexo = []
        #     cor = []
        #     endereco = []
        #     transporte = []
        #     nome_social = []
        #     possui_deficiencia = []
        #     deficiencia = []
        #     telefone = []
        #     celular = []
        #     email_aluno = []
        #     nome_responsavel = []
        #     numero_responsavel = []
        #     cpf_responsavel = []
        #     cpf_aluno = []
        #     bolsa_familia = []
        #     cns = []
        #     ra = []
        #     nivel = []
        #     etapa = []
        #     turno = []
        #     turma = []
        #     periodo_letivo = []
        #     data_matricula = []
        #     situacao_matricula = []
        #     ano_referencia = []
        #     exclusivo_aee = []

        #     dados = pd.read_excel('C:/Users/franklin.farias/Documents/Programação/HUMBERTOSOARES.xlsx', dtype=object)

        #     x = dados['MUNICIPIO'].tolist()
        #     foritem(x, municipio)
        #     x = dados['ESCOLA'].tolist()
        #     foritem(x, escola)
        #     x = dados['CD_CENSO'].tolist()
        #     foritem(x, inep_escola)
        #     x = dados['NM_ALUNO'].tolist()
        #     foritem(x, nome_aluno)
        #     x = dados['CD_INEP'].tolist()
        #     foritem(x, inep_aluno)
        #     x = dados['DT_NASCIMENTO'].tolist()
        #     foritem(x, nascimento)
        #     x = dados['NM_PAI'].tolist()
        #     foritem(x, nome_pai)
        #     x = dados['NM_MAE'].tolist()
        #     foritem(x, nome_mae)
        #     x = dados['TP_SEXO'].tolist()
        #     foritem(x, sexo)
        #     x = dados['TP_COR'].tolist()
        #     foritem(x, cor)
        #     x = dados['ENDERECO'].tolist()
        #     foritem(x, endereco)
        #     x = dados['TRANSPORTE_ESCOLAR'].tolist()
        #     foritem(x, transporte)
        #     x = dados['NM_ALUNO_SOCIAL'].tolist()
        #     foritem(x, nome_social)
        #     x = dados['POSSUI_DEFICIENCIA'].tolist()
        #     foritem(x, possui_deficiencia)
        #     x = dados['DC_DEFICIENCIA'].tolist()
        #     foritem(x, deficiencia)
        #     x = dados['NU_TELEFONE'].tolist()
        #     foritem(x, telefone)
        #     x = dados['NU_TELEFONE_CELULAR'].tolist()
        #     foritem(x, celular)
        #     x = dados['ED_EMAIL_ALUNO'].tolist()
        #     foritem(x, email_aluno)
        #     x = dados['NM_RESPONSAVEL'].tolist()
        #     foritem(x, nome_responsavel)
        #     x = dados['NU_RESPONSAVEL_TELEFONE'].tolist()
        #     foritem(x, numero_responsavel)
        #     x = dados['NU_RESPONSAVEL_CPF'].tolist()
        #     foritem(x, cpf_responsavel)
        #     x = dados['NU_CPF'].tolist()
        #     foritem(x, cpf_aluno)
        #     x = dados['BOLSA_FAMILIA'].tolist()
        #     foritem(x, bolsa_familia)
        #     x = dados['NU_CNS'].tolist()
        #     foritem(x, cns)
        #     x = dados['RA'].tolist()
        #     foritem(x, ra)
        #     x = dados['DC_NIVEL'].tolist()
        #     foritem(x, nivel)
        #     x = dados['DC_ETAPA'].tolist()
        #     foritem(x, etapa)
        #     x = dados['DC_TURNO'].tolist()
        #     foritem(x, turno)
        #     x = dados['DC_TURMA'].tolist()
        #     foritem(x, turma)
        #     x = dados['DC_PERIODO'].tolist()
        #     foritem(x, periodo_letivo)
        #     x = dados['DATA_MATRICULA'].tolist()
        #     foritem(x, data_matricula)
        #     x = dados['SITUACAO_MATRICULA'].tolist()
        #     foritem(x, situacao_matricula)
        #     x = dados['ANO_REFERENCIA'].tolist()
        #     foritem(x, ano_referencia)
        #     x = dados['EXCLUSIVO_AEE'].tolist()
        #     foritem(x, exclusivo_aee)


        #     for item in range(len(inep_escola)):
        #         simaed = Aux_simaed()
        #         simaed.municipio = municipio[item]
        #         simaed.escola = escola[item]
        #         simaed.inep_escola = inep_escola[item]
        #         simaed.nome_aluno = nome_aluno[item]
        #         simaed.inep_aluno = inep_aluno[item]
        #         simaed.nascimento = nascimento[item]
        #         simaed.nome_pai = nome_pai[item]
        #         simaed.nome_mae = nome_mae[item]
        #         simaed.sexo = sexo[item]
        #         simaed.cor = cor[item]
        #         simaed.endereco = endereco[item]
        #         simaed.transporte = transporte[item]
        #         simaed.nome_social = nome_social[item]
        #         simaed.possui_deficiencia = possui_deficiencia[item]
        #         simaed.deficiencia = deficiencia[item]
        #         simaed.telefone = telefone[item]
        #         simaed.celular = celular[item]
        #         simaed.email_aluno = email_aluno[item]
        #         simaed.nome_responsavel = nome_responsavel[item]
        #         simaed.numero_responsavel = numero_responsavel[item]
        #         simaed.cpf_responsavel = cpf_responsavel[item]
        #         simaed.cpf_aluno = cpf_aluno[item]
        #         simaed.bolsa_familia = bolsa_familia[item]
        #         simaed.cns = cns[item]
        #         simaed.ra = ra[item]
        #         simaed.nivel = nivel[item]
        #         simaed.etapa = etapa[item]
        #         simaed.turno = turno[item]
        #         simaed.turma = turma[item]
        #         simaed.periodo_letivo = periodo_letivo[item]
        #         simaed.data_matricula = data_matricula[item]
        #         simaed.situacao_matricula = situacao_matricula[item]
        #         simaed.ano_referencia = ano_referencia[item]
        #         simaed.exclusivo_aee = exclusivo_aee[item]
        #         simaed.save()

        # teste = usuarios()
        return HttpResponseRedirect('/atena')

        # PDF PARA GERAR LISTA DE TODOS OS SERVIDORES QUE REALIZARAM A ATUALIZAÇÃO DE LOTACAO
        # template_path = 'atena/teste-atualizacao-pdf.html'

        # url_logo = str(settings.BASE_DIR) + '/static/assets/img/detei.png'
        # url_atena = str(settings.BASE_DIR) + '/static/assets/img/ATENA-AZUL.svg'

        # atualizacoes = Confirmacao_lotacao.objects.all()

        # contexto = {'atualizacoes': atualizacoes,'url_logo':url_logo, 'url_atena':url_atena}

        # response = HttpResponse(content_type='application/pdf')
        # response['Content-Disposition'] = 'attachment; filename="Atualização de Lotação - GERAL.pdf"'

        # template = get_template(template_path)
        # html = template.render(contexto)

        # pisa_status = pisa.CreatePDF(html, dest=response)
        # if pisa_status.err:
        #     return HttpResponse('We had some errpors <pre>' + html + '</pre>')
        # return response


        # BIBLIOTECA PARA INSERIR DADOS NO RELATORIO DO SIMAED
        # # Pega o caminho do arquivo até a pasta do projeto (nl_SEE)
        # urls = str(settings.BASE_DIR)

        # # Abre o arquivo pelo caminho do documento.
        # wb = load_workbook(filename= urls + '/aplicacoes/dinem/relatorios/Pasta1.xlsx')

        # # Identifaca a aba da planilha pelo nome.
        # sh = wb['Plan1']

        # # Dados que serão inseridos na planilha.
        # nome = 'Franklin Farias'
        # idade = '22'
        # sexo = 'M'

        # # Instancia a variavel indicando a celula que vai receber o valor, com a posição de linha e coluna.
        # # A variavel recebe o valor dessa celula substituindo o nome que indicado na celula pelo valor desejado.
        # sh.cell(row= 1, column= 2).value = (sh.cell(row= 1, column= 2).value).replace('nome.pessoa', nome)
        # sh.cell(row= 2, column= 2).value = (sh.cell(row= 2, column= 2).value).replace('idade.pessoa', idade)
        # sh.cell(row= 3, column= 2).value = (sh.cell(row= 3, column= 2).value).replace('sexo.pessoa', sexo)

        # # Configuração indicando que o tipo do arquivo (EXCEL) e o nome do documento.
        # response = HttpResponse(content_type='application/vnd.ms-excel')
        # response['Content-Disposition'] = 'attachment; filename="Teste.xlsx"'

        # # Salva os dados na planilha.
        # wb.save(response)

        # return response

    # ÉRICKEEEEEEEEEEEEEEEE
    # if user == 'erick.nascimento':
        # url = f'{str(settings.BASE_DIR)}/aplicacoes/atena/simaed/turmas_simaed.csv'

        # dados = pd.read_csv(url, nrows=100)

        # nomes = dados['NM_ALUNO'].tolist()
        # nomes_sociais = dados['NM_ALUNO_SOCIAL'].tolist()
        # nascimentos = dados['NM_ALUNO_SOCIAL'].tolist()
        # cpfs = dados['NU_CPF'].tolist()

        # for nome in nomes:
        #     print(nome)

        # SCRIPT PARA MIGRAR OS DADOS DAS DEPENDENCIAS
        # dependencias = Infraestrutura_dependencia.objects.all()

        # for dependencia in dependencias:
        #     if dependencia.tipo:
        #         print(f'{dependencia.descricao} | {dependencia.tipo.tipo} | {dependencia.tipo}')

        #         tipo_antigo = dependencia.tipo.tipo
        #         print(tipo_antigo)

        #         try:
        #             tipo_novo = Infraestrutura_dependencia_tipo.objects.get(tipo= tipo_antigo)
        #             print(tipo_novo)

        #             dependencia.tipo = tipo_novo

        #             # dependencia.save()
        #         except:
        #             print('tipo não encontrato')

        #         print()
        # return TemplateResponse(request, template_name, locals())

    if user == 'vitor.cunha':
        # LEITURA DA PLANILHA E ALINHAMENTO DOS DADOS NO ARRAYS
        dados = pd.read_excel('D:/Users/vitor.cunha/OneDrive/Documents/turmalina.xls', dtype=object, sheet_name='b')

        escolas = dados['NOMEFORMATADO'].tolist()
        turmalina = dados['turmalina'].tolist()
        municipios = dados['MUNICÍPIOS'].tolist()

        # Model
        dado_Escola = Endereco.objects.all()

        escola_formatada = []

        for row in dado_Escola:
            escola_formatada.append([row.escola.id,row.escola.nome_escola,row.municipio.replace('ã','a').replace('á','a').replace('â','a').replace('ó','o').replace('é','e').upper()])

        contador = 0

        for rowExcel in range(len(dados)):
            for rowBanco in range(len(escola_formatada)):
                if (escolas[rowExcel] in escola_formatada[rowBanco][1]) and (municipios[rowExcel] == escola_formatada[rowBanco][2]):
                    Escola.objects.filter(id=escola_formatada[rowBanco][0]).update(
                        cod_turmalina=turmalina[rowExcel]
                    )
                    contador += 1

        return TemplateResponse(request, template_name, locals())