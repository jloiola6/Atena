from django.conf import settings
from django.http import HttpResponse
from openpyxl import *

from aplicacoes.administracao.models import *
from aplicacoes.dinem.models import *

def cabecalho_por_pagina(l, sh):
    #Linha 1 Cabeçalho
        sh.cell(row= l, column= 1).value = sh.cell(row= 4, column= 1).value
        sh.cell(row= l, column= 1).font = styles.Font(bold=True, name= 'Calibri', size= 12)
        sh.cell(row= l, column= 1).alignment = styles.Alignment(wrap_text= True, horizontal='center', vertical= 'center')

        sh.cell(row= l, column= 5).value = sh.cell(row= 4, column= 5).value
        sh.cell(row= l, column= 5).font = styles.Font(bold=True, name= 'Calibri', size= 12)
        sh.cell(row= l, column= 5).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= l, column= 17).value = sh.cell(row= 4, column= 17).value
        sh.cell(row= l, column= 17).font = styles.Font(bold=True, name= 'Calibri', size= 12)
        sh.cell(row= l, column= 17).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= l, column= 26).value = sh.cell(row= 4, column= 26).value
        sh.cell(row= l, column= 26).font = styles.Font(bold=True, name= 'Calibri', size= 12)
        sh.cell(row= l, column= 26).alignment = styles.Alignment(textRotation=90, horizontal='center', vertical= 'center')
        sh.row_dimensions[l].height = 15
        l += 1

        #Linha 2 Cabeçalho
        sh.cell(row= l, column= 1).value = sh.cell(row= 5, column= 1).value
        sh.cell(row= l, column= 17).value = sh.cell(row= 5, column= 17).value
        sh.cell(row= l, column= 17).font = styles.Font(bold=True, name= 'Calibri', size= 12)
        sh.cell(row= l, column= 17).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= l, column= 21).value = sh.cell(row= 5, column= 21).value
        sh.cell(row= l, column= 21).font = styles.Font(bold=True, name= 'Calibri', size= 12)
        sh.cell(row= l, column= 21).alignment = styles.Alignment(wrap_text= True, horizontal='center', vertical= 'center')

        sh.cell(row= l, column= 25).value = sh.cell(row= 5, column= 25).value
        sh.cell(row= l, column= 25).font = styles.Font(bold=True, name= 'Calibri', size= 12)
        sh.cell(row= l, column= 25).alignment = styles.Alignment(wrap_text= True, horizontal='center', vertical= 'center')
        sh.row_dimensions[l].height = 29
        l += 1

        #Linha 3 Cabeçalho
        sh.cell(row= l, column= 1).value = sh.cell(row= 6, column= 1).value
        for i in range(5, 26, 1):
            sh.cell(row= l, column= i).value = sh.cell(row= 6, column= i).value
            sh.cell(row= l, column= i).font = styles.Font(name= 'Calibri', size= 12)
            sh.cell(row= l, column= i).alignment = styles.Alignment(textRotation=90, wrap_text= True, horizontal='center', vertical= 'center')

        sh.row_dimensions[l].height = 110

        #Estrutura Curricular
        sh.merge_cells(f'A{l-2}:D{l}')
        #Formação Básica
        sh.merge_cells(f'E{l-2}:P{l-1}')
        #Diversificada
        sh.merge_cells(f'Q{l-1}:T{l-1}')
        #Itinerarios
        sh.merge_cells(f'Q{l-2}:Y{l-2}')
        #Formação Tecnica
        sh.merge_cells(f'Y{l-1}:Y{l-1}')
        #Resultado
        sh.merge_cells(f'Z{l-2}:Z{l}')
        #Rotas
        sh.merge_cells(f'U{l-1}:X{l-1}')

        l += 1

def turma_parcial_terceiro(request):
    turma = Turmas.objects.get(id= request.GET.get('id'))

    endereco = turma.endereco
    contato = Contato.objects.filter(endereco= endereco, tipo_contato__in=['T', 'C'])
    email = Contato.objects.filter(endereco=endereco, tipo_contato='E')

    aluno_turma = Aluno_turma.objects.filter(turma= turma)
    id_aluno = []
    for item in aluno_turma:
        id_aluno.append(item.aluno.id)

    relatorios = RelatorioFinal.objects.filter(aluno__in= id_aluno, etapa= '3ª Série').order_by('aluno__nome')

    wb = load_workbook(filename= str(settings.BASE_DIR) + '/aplicacoes/dinem/relatorios/ano_letivo_2022/turma_3_serie/turma_parcial_terceiro_2022_novo.xlsx')
    sh = wb['Relatório Final-Parcial-3ª']

    def cabecalho():
        sh.cell(row= l, column= 1).value = sh.cell(row= 4, column= 1)

    #Dados da escola
    sh.cell(row= 1, column= 11).value = (sh.cell(row= 1, column= 11).value).replace('NOME DA ESCOLA', endereco.escola.nome_escola)
    sh.cell(row= 1, column= 11).font = styles.Font(name= 'Calibri', size= 14)
    sh.cell(row= 1, column= 11).value = (sh.cell(row= 1, column= 11).value).replace('Endereco', endereco.rua)
    if endereco.cep != '':
        sh.cell(row= 1, column= 11).value = (sh.cell(row= 1, column= 11).value).replace('CEP', endereco.cep)
    else:
        sh.cell(row= 1, column= 11).value = (sh.cell(row= 1, column= 11).value).replace('-CEP', '')

    sh.cell(row= 1, column= 11).value = (sh.cell(row= 1, column= 11).value).replace('Municipio', endereco.municipio)
    sh.cell(row= 1, column= 11).value = (sh.cell(row= 1, column= 11).value).replace('Telefone', contato[0].contato)
    sh.cell(row= 1, column= 11).value = (sh.cell(row= 1, column= 11).value).replace('email', email[0].contato)

    #Dados da turma
    sh.cell(row= 3, column= 5).value = (sh.cell(row= 3, column= 5).value).replace('variavel_serie', '3ª Série')
    sh.cell(row= 3, column= 10).value = (sh.cell(row= 3, column= 10).value).replace('variavel_turma', turma.nome)
    sh.cell(row= 3, column= 17).value = (sh.cell(row= 3, column= 17).value).replace('variavel_turno', turma.turno)

    # #Dados do Aluno
    sh.row_dimensions[8].height = 20

    paginas = False

    l = 8
    ordem = contador = 1
    for relatorio in relatorios:
        sh.cell(row= l, column= 1).value = ordem

        sh.row_dimensions[l].height = 20
        if ordem < 10:
            sh.cell(row= l, column= 1).value = f'0{ordem}'
        else:
            sh.cell(row= l, column= 1).value = str(ordem)
        sh.merge_cells(f'B{l}:D{l}')
        sh.cell(row= l, column= 2).value = relatorio.aluno.nome

        if relatorio.resultado == 'TRANSFERIDO':
            sh.merge_cells(f'E{l}:Z{l}')
            sh.cell(row= l, column= 5).value = 'TRANSFERIDO'
            sh.cell(row= l, column= 5).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        elif relatorio.resultado == 'DESISTENTE':
            sh.merge_cells(f'E{l}:Z{l}')
            sh.cell(row= l, column= 5).value = 'DESISTENTE'
            sh.cell(row= l, column= 5).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        else:
            sh.cell(row= l, column= 5).value = 'Média'
            sh.cell(row= l, column= 6).value = relatorio.portugues
            sh.cell(row= l, column= 7).value = '-'
            sh.cell(row= l, column= 8).value = '-'
            sh.cell(row= l, column= 9).value = relatorio.ed_fisica
            sh.cell(row= l, column= 10).value = relatorio.matematica
            sh.cell(row= l, column= 11).value = '-'
            sh.cell(row= l, column= 12).value = '-'
            sh.cell(row= l, column= 13).value = '-'
            sh.cell(row= l, column= 14).value = '-'
            sh.cell(row= l, column= 15).value = '-'
            sh.cell(row= l, column= 16).value = '-'
            sh.cell(row= l, column= 17).value = '-'
            sh.cell(row= l, column= 18).value = relatorio.projeto_vida
            sh.cell(row= l, column= 19).value = '-'
            sh.cell(row= l, column= 20).value = relatorio.pos_medio
            sh.cell(row= l, column= 21).value = relatorio.linguagens_tecnologias
            sh.cell(row= l, column= 22).value = relatorio.ciencias_humanas
            sh.cell(row= l, column= 23).value = relatorio.matematica_tecnologias
            sh.cell(row= l, column= 24).value = relatorio.ciencias_natureza
            sh.cell(row= l, column= 25).value = relatorio.formacao_tecnica

            if relatorio.resultado == 'APROVADO':
                resultado = 'AP'
            elif relatorio.resultado == 'APROVADO COM DEPENDÊNCIA':
                resultado = 'APD'
            elif relatorio.resultado == 'REPROVADO':
                resultado = 'RP'
            sh.cell(row= l, column= 26).value = resultado

        #Centralizando células
        sh.cell(row= l, column= 1).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 2).alignment = styles.Alignment(vertical= 'center')
        sh.cell(row= l, column= 5).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 6).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 8).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 10).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 12).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 14).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 15).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 16).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 17).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 19).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 20).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 21).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 22).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 23).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 24).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 25).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= l, column= 26).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        l += 1
        ordem += 1
        contador += 1

        if contador == 35 and not paginas:
            paginas = True
            print(ordem)
            cabecalho_por_pagina(l, sh)
            l += 3
            contador = 1
            continue
        elif contador == 36 and paginas == True:
            cabecalho_por_pagina(l, sh)
            l += 3
            contador = 1
            continue

    #Detalhes da Planilha
    sh2 = wb['Plan1']

    sh.cell(row= l, column= 1).value = sh2.cell(row= 5, column= 30).value
    sh.cell(row= l, column= 1).alignment = styles.Alignment(wrap_text= True, horizontal='left', vertical= 'top')
    sh.merge_cells(f'A{l}:D{l}')

    sh.cell(row= l, column= 5).value = sh2.cell(row= 5, column= 34).value
    sh.cell(row= l, column= 5).alignment = styles.Alignment(wrap_text= True, horizontal='left', vertical= 'top')
    sh.merge_cells(f'E{l}:Z{l}')


    # #Criando Rodapé
    sh.cell(row= l+1, column= 1).value = sh2.cell(row= 1, column= 1).value
    sh.cell(row= l+1, column= 1).alignment = styles.Alignment(horizontal='left', vertical= 'top')
    sh.merge_cells(f'A{l+1}:Z{l+1}')

    sh.cell(row= l+2, column= 1).value = sh2.cell(row= 2, column= 1).value
    sh.cell(row= l+2, column= 1).alignment = styles.Alignment(horizontal='left', vertical= 'top')
    sh.merge_cells(f'A{l+2}:F{l+2}')

    sh.cell(row= l+2, column= 7).value = sh2.cell(row= 2, column= 7).value
    sh.cell(row= l+2, column= 7).alignment = styles.Alignment(horizontal='left', vertical= 'top')
    sh.merge_cells(f'G{l+2}:Q{l+2}')

    sh.cell(row= l+2, column= 18).value = sh2.cell(row= 2, column= 18).value
    sh.cell(row= l+2, column= 18).alignment = styles.Alignment(horizontal='left', vertical= 'top')
    sh.merge_cells(f'R{l+2}:Z{l+2}')


    #Definindo altura das linhas do Rodapé
    sh.row_dimensions[l].height = 60
    sh.row_dimensions[l+1].height = 114
    sh.row_dimensions[l+2].height = 64

    #Adicionando Bordas no rodapé
    thin = styles.Side(border_style="thin", color="000000")
    double = styles.Side(border_style="thin", color="000000")
    for i in range(8, l+3, 1):
        sh.cell(row= i, column= 1).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 2).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 3).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 4).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 5).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 6).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 7).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 8).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 9).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 10).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 11).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 12).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 13).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 14).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 15).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 16).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 17).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 18).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 19).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 20).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 21).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 22).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 23).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 24).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 25).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 26).border = styles.Border(top=double, left=thin, right=thin, bottom=double)

    wb.remove(wb['Plan1'])

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Turma Parcial 3ª Série.xlsx"'

    wb.save(response)

    return response