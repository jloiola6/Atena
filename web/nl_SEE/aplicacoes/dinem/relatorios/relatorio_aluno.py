from django.http import HttpResponse
from openpyxl import *
from openpyxl.formula.translate import Translator

from aplicacoes.administracao.models import Endereco, Escola, Contato, Turmas, Grade, Gestor_Escolar, Aluno_turma
from aplicacoes.dinem.models import RelatorioFinal, Aluno
from django.conf import settings

def zerando_horas(serie, sh):
    total1 = str(sh.cell(row= 23, column= 18).value)
    total2 = str(sh.cell(row= 28, column= 18).value)
    if '.4' in total1:
        final = '.4'
        try:
            total1 = int(total1[0:4])
        except:
            total1 = int(total1[0:3])

        total2 = int(total2[0:3])

    if serie == 1:
        coluna = 12
        sh.cell(row= 23, column= 18).value = str(total1 - 840) + final
        sh.cell(row= 23, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        total = float(sh.cell(row= 28, column= 18).value)
        sh.cell(row= 28, column= 18).value = str(total2 - 200) + final
        sh.cell(row= 28, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    elif serie == 2:
        coluna = 14
        sh.cell(row= 23, column= 18).value = str(total1 - 646)
        sh.cell(row= 23, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= 28, column= 18).value = str(total2 - 154)
        sh.cell(row= 28, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    else:
        coluna = 16
        try:
            sh.cell(row= 23, column= 18).value = str(total1 - 160) + final
            sh.cell(row= 23, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

            sh.cell(row= 28, column= 18).value = str(total2 - 64) + final
            sh.cell(row= 28, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        except:
            pass

    for linha in range(11, 29, 1):
        try:
            if linha not in (23, 28):
                # hora = sh.cell(row= linha, column= coluna).value
                # sh.cell(row= linha, column= coluna).alignment = styles.Alignment(horizontal='center', vertical= 'center')

                sh.cell(row= linha, column= coluna).value = '-'
                sh.cell(row= linha, column= coluna).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            else:
                sh.cell(row= linha, column= coluna-1).value = '-'
                sh.cell(row= linha, column= coluna-1).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        except:
            pass


def relatorio_aluno(request, id_aluno_turma):
    wb = load_workbook(filename= str(settings.BASE_DIR) + '/aplicacoes/dinem/relatorios/relatorio_aluno.xlsx')
    sh = wb['Table 1']

    aluno_turma = Aluno_turma.objects.get(id= id_aluno_turma)
    aluno = aluno_turma.aluno
    turma = aluno_turma.turma
    escola = aluno_turma.turma.endereco
    endereco = Endereco.objects.get(id= escola.id)
    contato = Contato.objects.filter(endereco= endereco)

    relatorio1 = RelatorioFinal.objects.get(aluno= aluno, etapa= '1ª Série')
    relatorio2 = RelatorioFinal.objects.get(aluno= aluno, etapa= '2ª Série')
    relatorio3 = RelatorioFinal.objects.get(aluno= aluno, etapa= '3ª Série')

    #Dados da escola
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('NOME DA ESCOLA', escola.escola.nome_escola)
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('Endereço', endereco.rua)
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('CEP', endereco.cep)
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('Município', endereco.municipio)
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('Telefone', contato[0].contato)
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('e-mail', contato[1].contato)

    sh.cell(row= 2, column= 1).font = styles.Font(name= 'Carlito', size= 8)

    resultados = []
    serie = 1
    for relatorio in [relatorio1.resultado, relatorio2.resultado, relatorio3.resultado]:
        if relatorio == 'APROVADO':
            resultado = 'AP'
        elif relatorio == 'APROVADO COM DEPENDÊNCIA':
            resultado = 'APD'
        elif relatorio == 'REPROVADO':
            resultado = 'RP'
        elif relatorio == 'Desistente':
            resultado = 'Desistente'
            zerando_horas(serie, sh)
        else:
            resultado = 'Transferido'
            zerando_horas(serie, sh)
        resultados.append(resultado)
        serie += 1


    #Dados do Aluno
    sh.cell(row= 5, column= 1).value = (sh.cell(row= 5, column= 1).value).replace('aluno.nome', aluno.nome)
    sh.cell(row= 5, column= 1).font = styles.Font(name= 'Arial', size= 7)

    data_nascimento = aluno.nascimento.split('-')
    data_nascimento = f'{data_nascimento[2]}/{data_nascimento[1]}/{data_nascimento[0]}'

    sh.cell(row= 6, column= 1).value = (sh.cell(row= 6, column= 1).value).replace('aluno.data', data_nascimento)
    sh.cell(row= 6, column= 1).font = styles.Font(name= 'Arial', size= 7)

    sh.cell(row= 6, column= 11).value = (sh.cell(row= 6, column= 11).value).replace('aluno.nacionalidade', aluno.naturalidade.upper())
    sh.cell(row= 6, column= 11).font = styles.Font(name= 'Arial', size= 7)

    sh.cell(row= 6, column= 17).value = (sh.cell(row= 6, column= 17).value).replace('aluno.uf', aluno.uf.upper())
    sh.cell(row= 6, column= 17).font = styles.Font(name= 'Arial', size= 7)

    sh.cell(row= 7, column= 4).value = (sh.cell(row= 7, column= 4).value).replace('aluno.pai', aluno.nome_pai.upper())
    sh.cell(row= 7, column= 4).font = styles.Font(name= 'Arial', size= 7)

    sh.cell(row= 8, column= 4).value = (sh.cell(row= 8, column= 4).value).replace('aluno.mae', aluno.nome_mae.upper())
    sh.cell(row= 8, column= 4).font = styles.Font(name= 'Arial', size= 7)


    # Dados relatorio1
    sh.cell(row= 11, column= 11).value = (sh.cell(row= 11, column= 11).value).replace('relatorio1.portugues', relatorio1.portugues)
    sh.cell(row= 11, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 12, column= 11).value = (sh.cell(row= 12, column= 11).value).replace('relatorio1.arte', relatorio1.arte)
    sh.cell(row= 12, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 13, column= 11).value = (sh.cell(row= 13, column= 11).value).replace('relatorio1.ingles', relatorio1.ingles)
    sh.cell(row= 13, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 14, column= 11).value = (sh.cell(row= 14, column= 11).value).replace('relatorio1.ed_fisica', relatorio1.ed_fisica)
    sh.cell(row= 14, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 15, column= 11).value = (sh.cell(row= 15, column= 11).value).replace('relatorio1.matematica', relatorio1.matematica)
    sh.cell(row= 15, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 16, column= 11).value = (sh.cell(row= 16, column= 11).value).replace('relatorio1.fisica', relatorio1.fisica)
    sh.cell(row= 16, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 17, column= 11).value = (sh.cell(row= 17, column= 11).value).replace('relatorio1.quimica', relatorio1.quimica)
    sh.cell(row= 17, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 18, column= 11).value = (sh.cell(row= 18, column= 11).value).replace('relatorio1.biologia', relatorio1.biologia)
    sh.cell(row= 18, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 19, column= 11).value = (sh.cell(row= 19, column= 11).value).replace('relatorio1.historia', relatorio1.historia)
    sh.cell(row= 19, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 20, column= 11).value = (sh.cell(row= 20, column= 11).value).replace('relatorio1.geografia', relatorio1.geografia)
    sh.cell(row= 20, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 21, column= 11).value = (sh.cell(row= 21, column= 11).value).replace('relatorio1.filosofia', relatorio1.filosofia)
    sh.cell(row= 21, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 22, column= 11).value = (sh.cell(row= 22, column= 11).value).replace('relatorio1.sociologia', relatorio1.sociologia)
    sh.cell(row= 22, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 24, column= 11).value = (sh.cell(row= 24, column= 11).value).replace('relatorio1.espanhol', relatorio1.espanhol)
    sh.cell(row= 24, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 25, column= 11).value = (sh.cell(row= 25, column= 11).value).replace('relatorio1.pj_vd', relatorio1.projeto_vida)
    sh.cell(row= 25, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 26, column= 11).value = (sh.cell(row= 26, column= 11).value).replace('relatorio1.lt_ch', relatorio1.lt_ch)
    sh.cell(row= 26, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 27, column= 11).value = (sh.cell(row= 27, column= 11).value).replace('relatorio1.mt_ch', relatorio1.mt_cn)
    sh.cell(row= 27, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')


    # Dados relatorio2
    sh.cell(row= 11, column= 13).value = (sh.cell(row= 11, column= 13).value).replace('relatorio2.portugues', relatorio2.portugues)
    sh.cell(row= 11, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 12, column= 13).value = (sh.cell(row= 12, column= 13).value).replace('relatorio2.arte', relatorio2.arte)
    sh.cell(row= 12, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 13, column= 13).value = (sh.cell(row= 13, column= 13).value).replace('relatorio2.ingles', relatorio2.ingles)
    sh.cell(row= 13, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 14, column= 13).value = (sh.cell(row= 14, column= 13).value).replace('relatorio2.ed_fisica', relatorio2.ed_fisica)
    sh.cell(row= 14, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 15, column= 13).value = (sh.cell(row= 15, column= 13).value).replace('relatorio2.matematica', relatorio2.matematica)
    sh.cell(row= 15, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 16, column= 13).value = (sh.cell(row= 16, column= 13).value).replace('relatorio2.fisica', relatorio2.fisica)
    sh.cell(row= 16, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 17, column= 13).value = (sh.cell(row= 17, column= 13).value).replace('relatorio2.quimica', relatorio2.quimica)
    sh.cell(row= 17, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 18, column= 13).value = (sh.cell(row= 18, column= 13).value).replace('relatorio2.biologia', relatorio2.biologia)
    sh.cell(row= 18, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 19, column= 13).value = (sh.cell(row= 19, column= 13).value).replace('relatorio2.historia', relatorio2.historia)
    sh.cell(row= 19, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 20, column= 13).value = (sh.cell(row= 20, column= 13).value).replace('relatorio2.geografia', relatorio2.geografia)
    sh.cell(row= 20, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 21, column= 13).value = (sh.cell(row= 21, column= 13).value).replace('relatorio2.filosofia', relatorio2.filosofia)
    sh.cell(row= 21, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 22, column= 13).value = (sh.cell(row= 22, column= 13).value).replace('relatorio2.sociologia', relatorio2.sociologia)
    sh.cell(row= 22, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 24, column= 13).value = (sh.cell(row= 24, column= 13).value).replace('relatorio2.espanhol', relatorio2.espanhol)
    sh.cell(row= 24, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 25, column= 13).value = (sh.cell(row= 25, column= 13).value).replace('relatorio2.pj_vd', relatorio2.projeto_vida)
    sh.cell(row= 25, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 26, column= 13).value = (sh.cell(row= 26, column= 13).value).replace('relatorio2.lt_ch', relatorio2.lt_ch)
    sh.cell(row= 26, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 27, column= 13).value = (sh.cell(row= 27, column= 13).value).replace('relatorio2.mt_ch', relatorio2.mt_cn)
    sh.cell(row= 27, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')


    # Dados relatorio3
    sh.cell(row= 11, column= 15).value = (sh.cell(row= 11, column= 15).value).replace('relatorio3.portugues', relatorio3.portugues)
    sh.cell(row= 11, column= 15).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 15, column= 15).value = (sh.cell(row= 15, column= 15).value).replace('relatorio3.matematica', relatorio3.matematica)
    sh.cell(row= 15, column= 15).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 25, column= 15).value = (sh.cell(row= 25, column= 15).value).replace('relatorio3.pj_vd', relatorio3.projeto_vida)
    sh.cell(row= 25, column= 15).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 32, column= 16).value = (sh.cell(row= 32, column= 16).value).replace('relatorio3.IC', relatorio3.investigacao)
    sh.cell(row= 32, column= 16).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 33, column= 16).value = (sh.cell(row= 33, column= 16).value).replace('relatorio3.PC', relatorio3.criativos)
    sh.cell(row= 33, column= 16).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 34, column= 16).value = (sh.cell(row= 34, column= 16).value).replace('relatorio3.MIS', relatorio3.sociocultural)
    sh.cell(row= 34, column= 16).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 35, column= 16).value = (sh.cell(row= 35, column= 16).value).replace('relatorio3.EMP', relatorio3.empreendedorismo)
    sh.cell(row= 35, column= 16).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    if relatorio3.investigacao == '-':
        # total1 = '992'
        # total2 = '2638'
        total1 = '1.138'
        total2 = '2784'


        sh.cell(row= 34, column= 3).value = (sh.cell(row= 34, column= 3).value).replace('area_conhecimento', '-')
        sh.cell(row= 34, column= 3).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= 32, column= 18).value = '-'
        sh.cell(row= 32, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= 33, column= 18).value = '-'
        sh.cell(row= 33, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= 34, column= 18).value = '-'
        sh.cell(row= 34, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= 35, column= 18).value = '-'
        sh.cell(row= 35, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= 36, column= 18).value = '-'
        sh.cell(row= 36, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        # sh.cell(row= 38, column= 10).value = (sh.cell(row= 38, column= 10).value).replace('total1', total1)
        sh.cell(row= 38, column= 10).value = total1
        sh.cell(row= 38, column= 10).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        # sh.cell(row= 39, column= 10).value = (sh.cell(row= 39, column= 10).value).replace('total2', total2)
        sh.cell(row= 39, column= 10).value = total2
        sh.cell(row= 39, column= 10).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    else:
        # total1 = '1136'
        # total2 = '2782'
        total1 = '994'
        total2 = '2640'

        sh.cell(row= 34, column= 3).value = (sh.cell(row= 34, column= 3).value).replace('area_conhecimento', relatorio3.area_conhecimento)
        sh.cell(row= 34, column= 3).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= 34, column= 3).font = styles.Font(name= 'Carlito', size= 7, bold= True)

        sh.cell(row= 37, column= 18).value = '-'
        sh.cell(row= 37, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= 38, column= 10).value = (sh.cell(row= 38, column= 10).value).replace('total1', total1)
        sh.cell(row= 38, column= 10).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= 39, column= 10).value = (sh.cell(row= 39, column= 10).value).replace('total2', total2)
        sh.cell(row= 39, column= 10).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    c = 0
    for i in range(43, 46, 1):
        sh.cell(row= i, column= 7).value = (sh.cell(row= i, column= 7).value).replace('NOME DA ESCOLA', escola.escola.nome_escola)
        sh.cell(row= i, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= i, column= 7).font = styles.Font(name= 'Carlito', size= 9)

        # sh.cell(row= i, column= 6).value = (sh.cell(row= i, column= 6).value).replace('turma.turno', turma.turno)
        # sh.cell(row= i, column= 6).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        # sh.cell(row= i, column= 6).font = styles.Font(name= 'Carlito', size= 6)

        sh.cell(row= i, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= i, column= 18).value = (sh.cell(row= i, column= 18).value).replace(f'relatorio{c+1}.resultado', resultados[c])
        sh.cell(row= i, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        c += 1

    # celula = sh.active
    # sh['G42'] = '= SUM(1+1)'

    #Deletando Tab
    wb.remove(wb['Table 2'])

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Aluno.xlsx"'

    wb.save(response)

    return response
