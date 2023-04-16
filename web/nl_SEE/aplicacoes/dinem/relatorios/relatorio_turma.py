from django.http import HttpResponse
from openpyxl import *

from aplicacoes.administracao.models import Endereco, Escola, Contato, Turmas, Grade, Gestor_Escolar, Aluno_turma
from aplicacoes.dinem.models import RelatorioFinal
from django.conf import settings

def cabecalho_por_pagina(l, sh):
    #Linha 1 Cabeçalho
        sh.cell(row= l, column= 1).value = sh.cell(row= 4, column= 1).value
        sh.cell(row= l, column= 1).font = styles.Font(bold=True, name= 'Carlito', size= 9)
        sh.cell(row= l, column= 1).alignment = styles.Alignment(wrap_text= True, horizontal='left', vertical= 'top')

        sh.cell(row= l, column= 6).value = sh.cell(row= 4, column= 6).value
        sh.cell(row= l, column= 6).font = styles.Font(bold=True, name= 'Carlito', size= 9)
        sh.cell(row= l, column= 6).alignment = styles.Alignment(horizontal='center', vertical= 'top')

        sh.cell(row= l, column= 18).value = sh.cell(row= 4, column= 18).value
        sh.cell(row= l, column= 18).font = styles.Font(bold=True, name= 'Carlito', size= 9)
        sh.cell(row= l, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'top')

        sh.cell(row= l, column= 27).value = sh.cell(row= 4, column= 27).value
        sh.cell(row= l, column= 27).font = styles.Font(bold=True, name= 'Carlito', size= 9)
        sh.cell(row= l, column= 27).alignment = styles.Alignment(textRotation=90, horizontal='center', vertical= 'center')
        sh.row_dimensions[l].height = 15
        l += 1

        #Linha 2 Cabeçalho
        sh.cell(row= l, column= 1).value = sh.cell(row= 5, column= 1).value
        sh.cell(row= l, column= 18).value = sh.cell(row= 5, column= 18).value
        sh.cell(row= l, column= 18).font = styles.Font(bold=True, name= 'Carlito', size= 9)
        sh.cell(row= l, column= 18).alignment = styles.Alignment(horizontal='left', vertical= 'top')

        sh.cell(row= l, column= 22).value = sh.cell(row= 5, column= 22).value
        sh.cell(row= l, column= 22).font = styles.Font(bold=True, name= 'Carlito', size= 9)
        sh.cell(row= l, column= 22).alignment = styles.Alignment(wrap_text= True, horizontal='left', vertical= 'top')

        sh.cell(row= l, column= 26).value = sh.cell(row= 5, column= 26).value
        sh.cell(row= l, column= 26).font = styles.Font(bold=True, name= 'Carlito', size= 9)
        sh.cell(row= l, column= 26).alignment = styles.Alignment(textRotation=90, wrap_text= True)
        sh.row_dimensions[l].height = 29
        l += 1

        #Linha 3 Cabeçalho
        sh.cell(row= l, column= 1).value = sh.cell(row= 6, column= 1).value
        for i in range(6, 26, 1):
            sh.cell(row= l, column= i).value = sh.cell(row= 6, column= i).value
            sh.cell(row= l, column= i).font = styles.Font(bold=True, name= 'Carlito', size= 9)
            sh.cell(row= l, column= i).alignment = styles.Alignment(textRotation=90, wrap_text= True)

        sh.row_dimensions[l].height = 95

        sh.merge_cells(f'A{l-2}:E{l}')
        sh.merge_cells(f'F{l-2}:Q{l-1}')
        sh.merge_cells(f'R{l-1}:U{l-1}')
        sh.merge_cells(f'V{l-1}:Y{l-1}')
        sh.merge_cells(f'Z{l-1}:Z{l}')
        sh.merge_cells(f'AA{l-2}:AA{l}')
        sh.merge_cells(f'R{l-2}:Z{l-2}')

        l += 1

def relatorio_turma(request):
    turma = Turmas.objects.get(id= request.GET.get('id'))
    escola = turma.escola
    endereco = Endereco.objects.get(escola= escola)
    contato = Contato.objects.filter(endereco= endereco)
    aluno_turma = Aluno_turma.objects.filter(turma= turma)
    id_aluno = []
    for item in aluno_turma:
        id_aluno.append(item.aluno.id)
    relatorios = RelatorioFinal.objects.filter(aluno__in= id_aluno, etapa= '3ª Série').order_by('aluno__nome')

    wb = load_workbook(filename= str(settings.BASE_DIR) + '/aplicacoes/dinem/relatorios/relatorio_turma.xlsx')
    sh = wb['Table 1']

    def cabecalho():
        sh.cell(row= l, column= 1).value = sh.cell(row= 4, column= 1)

    #Dados da escola
    sh.cell(row= 1, column= 12).value = (sh.cell(row= 1, column= 12).value).replace('NOME DA ESCOLA', escola.nome_escola)
    sh.cell(row= 1, column= 12).value = (sh.cell(row= 1, column= 12).value).replace('Endereço', endereco.rua)
    if endereco.cep == '':
        sh.cell(row= 1, column= 12).value = (sh.cell(row= 1, column= 12).value).replace('CEP', endereco.cep)
    else:
        sh.cell(row= 1, column= 12).value = (sh.cell(row= 1, column= 12).value).replace('-CEP', '')

    sh.cell(row= 1, column= 12).value = (sh.cell(row= 1, column= 12).value).replace('Município', endereco.municipio)
    # sh.cell(row= 1, column= 12).value = (sh.cell(row= 1, column= 12).value).replace('Telefone', contato[0].contato)
    # sh.cell(row= 1, column= 12).value = (sh.cell(row= 1, column= 12).value).replace('Telefone', contato.email)

    #Dados da turma
    sh.cell(row= 3, column= 6).value = (sh.cell(row= 3, column= 6).value).replace('variavel_serie', '3ª Série')
    sh.cell(row= 3, column= 12).value = (sh.cell(row= 3, column= 12).value).replace('variavel_turma', turma.nome)
    sh.cell(row= 3, column= 16).value = (sh.cell(row= 3, column= 16).value).replace('variavel_turno', turma.turno)

    #Dados de Carga Horária
    sh.cell(row= 7, column= 7).value = '96h'
    sh.cell(row= 7, column= 11).value = '64h'
    sh.cell(row= 7, column= 19).value = '64h'
    sh.cell(row= 7, column= 22).value = '144h'
    sh.cell(row= 7, column= 23).value = '144h'
    sh.cell(row= 7, column= 24).value = '144h'
    sh.cell(row= 7, column= 25).value = '144h'
    sh.cell(row= 7, column= 26).value = '720h'

    sh.cell(row= 7, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    sh.cell(row= 7, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    sh.cell(row= 7, column= 19).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    sh.cell(row= 7, column= 22).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    sh.cell(row= 7, column= 23).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    sh.cell(row= 7, column= 24).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    sh.cell(row= 7, column= 25).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    sh.cell(row= 7, column= 26).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    #Dados do Aluno
    sh.row_dimensions[7].height = 14

    paginas = False

    l = 8
    ordem = contador = 1
    for relatorio in relatorios:
        sh.cell(row= l, column= 1).value = ordem

        sh.row_dimensions[l].height = 17
        if ordem < 10:
            sh.cell(row= l, column= 1).value = f'0{ordem}'
        else:
            sh.cell(row= l, column= 1).value = str(ordem)
        sh.merge_cells(f'B{l}:E{l}')
        sh.cell(row= l, column= 2).value = relatorio.aluno.nome

        if relatorio.resultado == 'TRANSFERIDO':
            # sh.merge_cells('F8:AA8')
            sh.merge_cells(f'F{l}:AA{l}')
            sh.cell(row= l, column= 6).value = 'TRANFERIDO'
            sh.cell(row= l, column= 6).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        elif relatorio.resultado == 'DESISTENTE':
            # sh.merge_cells('F8:AA8')
            sh.merge_cells(f'F{l}:AA{l}')
            sh.cell(row= l, column= 6).value = 'TRANFERIDO'
            sh.cell(row= l, column= 6).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        else:
            sh.cell(row= l, column= 6).value = 'Média'
            sh.cell(row= l, column= 7).value = relatorio.portugues
            sh.cell(row= l, column= 8).value = '-'
            sh.cell(row= l, column= 9).value = '-'
            sh.cell(row= l, column= 10).value = '-'
            sh.cell(row= l, column= 11).value = relatorio.matematica
            sh.cell(row= l, column= 12).value = '-'
            sh.cell(row= l, column= 13).value = '-'
            sh.cell(row= l, column= 14).value = '-'
            sh.cell(row= l, column= 15).value = '-'
            sh.cell(row= l, column= 16).value = '-'
            sh.cell(row= l, column= 17).value = '-'
            sh.cell(row= l, column= 18).value = '-'
            sh.cell(row= l, column= 19).value = relatorio.projeto_vida
            sh.cell(row= l, column= 20).value = '-'
            sh.cell(row= l, column= 21).value = '-'
            sh.cell(row= l, column= 22).value = relatorio.investigacao
            sh.cell(row= l, column= 23).value = relatorio.criativos
            sh.cell(row= l, column= 24).value = relatorio.sociocultural
            sh.cell(row= l, column= 25).value = relatorio.empreendedorismo
            if relatorio.investigacao == '-':
                sh.cell(row= l, column= 26).value = 'OP'
            else:
                sh.cell(row= l, column= 26).value = '-'

            if relatorio.resultado == 'APROVADO':
                resultado = 'AP'
            elif relatorio.resultado == 'APROVADO COM DEPENDÊNCIA':
                resultado = 'APD'
            elif relatorio.resultado == 'REPROVADO':
                resultado = 'RP'
            sh.cell(row= l, column= 27).value = resultado

            #Centralizando células
            sh.cell(row= l, column= 1).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 6).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 8).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 10).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 11).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 12).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 13).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 14).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 15).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 16).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 17).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 18).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 19).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 20).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 21).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 22).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 23).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 24).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 25).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 26).alignment = styles.Alignment(horizontal='center', vertical= 'center')
            sh.cell(row= l, column= 27).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        l += 1
        ordem += 1
        contador += 1

        if contador == 26 and not paginas:
            paginas = True
            cabecalho_por_pagina(l, sh)
            l += 3
            contador = 1
            continue
        elif contador == 33 and paginas == True:
            cabecalho_por_pagina(l, sh)
            l += 3
            contador = 1
            continue

    #Detalhes da Planilha
    sh2 = wb['Table 2']

    sh.cell(row= l, column= 1).value = sh2.cell(row= 5, column= 30).value
    sh.cell(row= l, column= 1).alignment = styles.Alignment(wrap_text= True, horizontal='left', vertical= 'top')
    sh.merge_cells(f'A{l}:D{l}')

    sh.cell(row= l, column= 5).value = sh2.cell(row= 5, column= 34).value
    sh.cell(row= l, column= 5).alignment = styles.Alignment(wrap_text= True, horizontal='left', vertical= 'top')
    sh.merge_cells(f'E{l}:AA{l}')


    # #Criando Rodapé
    sh.cell(row= l+1, column= 1).value = sh2.cell(row= 1, column= 1).value
    sh.cell(row= l+1, column= 1).alignment = styles.Alignment(horizontal='left', vertical= 'top')
    sh.merge_cells(f'A{l+1}:AA{l+1}')

    sh.cell(row= l+2, column= 1).value = sh2.cell(row= 2, column= 1).value
    sh.cell(row= l+2, column= 1).alignment = styles.Alignment(horizontal='left', vertical= 'top')
    sh.merge_cells(f'A{l+2}:F{l+2}')

    sh.cell(row= l+2, column= 7).value = sh2.cell(row= 2, column= 7).value
    sh.cell(row= l+2, column= 7).alignment = styles.Alignment(horizontal='left', vertical= 'top')
    sh.merge_cells(f'G{l+2}:Q{l+2}')

    sh.cell(row= l+2, column= 18).value = sh2.cell(row= 2, column= 18).value
    sh.cell(row= l+2, column= 18).alignment = styles.Alignment(horizontal='left', vertical= 'top')
    sh.merge_cells(f'R{l+2}:AA{l+2}')


    #Definindo altura das linhas do Rodapé
    sh.row_dimensions[l].height = 60
    sh.row_dimensions[l+1].height = 114
    sh.row_dimensions[l+2].height = 64



    #Adicionando Bordas no rodapé
    thin = styles.Side(border_style="thin", color="000000")
    double = styles.Side(border_style="thin", color="000000")
    for i in range(8, l+3, 1):
        sh.cell(row= i, column= 1).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 2).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 3).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 4).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 5).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 6).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 7).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 8).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 9).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 10).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 11).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 12).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 13).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 14).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 15).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 16).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 17).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 18).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 19).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 20).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 21).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 22).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 23).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 24).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 25).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 26).border = styles.Border(top=double, left=thin, right=thin, bottom=double)
        sh.cell(row= i, column= 27).border = styles.Border(top=double, left=thin, right=thin, bottom=double)


    #Deletando Tab
    wb.remove(wb['Table 2'])


    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Turma.xlsx"'

    wb.save(response)

    return response