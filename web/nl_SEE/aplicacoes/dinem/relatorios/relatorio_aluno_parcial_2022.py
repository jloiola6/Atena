from django.http import HttpResponse
from openpyxl import *
from openpyxl.formula.translate import Translator

from aplicacoes.administracao.models import Endereco, Escola, Contato, Turmas, Grade, Gestor_Escolar, Aluno_turma
from aplicacoes.dinem.models import RelatorioFinal, Aluno
from django.conf import settings



def dados_aluno_parcial(request, id_aluno_turma):
    urls = str(settings.BASE_DIR)

    # Abre o arquivo pelo caminho do documento.
    wb = load_workbook(filename= urls + '/aplicacoes/dinem/relatorios/relatorio_aluno_parcial_2022.xlsx')

    # Identifaca a aba da planilha pelo nome.
    sh = wb['Planilha1']

    # Dados que serão inseridos na planilha.
    aluno_turma = Aluno_turma.objects.get(id= id_aluno_turma)
    aluno = aluno_turma.aluno
    turma = aluno_turma.turma
    escola = aluno_turma.turma.endereco
    endereco = Endereco.objects.get(id= escola.id)
    email = Contato.objects.filter(endereco= endereco, tipo_contato = 'E')
    contato = Contato.objects.filter(endereco= endereco, tipo_contato__in = ['T', 'C'])

    # Dados que serão inseridos na planilha.
    relatorio1 = RelatorioFinal.objects.get(aluno= aluno, etapa= '1ª Série', turma= turma)
    relatorio2 = RelatorioFinal.objects.get(aluno= aluno, etapa= '2ª Série', turma= turma)
    relatorio3 = RelatorioFinal.objects.get(aluno= aluno, etapa= '3ª Série', turma= turma)

    resultados = []
    serie = 1
    for relatorio in [relatorio1.resultado, relatorio2.resultado, relatorio3.resultado]:
        if relatorio == 'APROVADO':
            resultado = 'AP'
        elif relatorio == 'APROVADO COM DEPENDÊNCIA':
            resultado = 'APD'
        elif relatorio == 'REPROVADO':
            resultado = 'RP'
        elif relatorio == 'Desistente':
            resultado = 'Desistente'
        else:
            resultado = 'Transferido'
        resultados.append(resultado)
        serie += 1


    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('NOME DA ESCOLA', escola.escola.nome_escola)
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('Endereço', endereco.rua)
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('CEP', endereco.cep)
    sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('Município', endereco.municipio)

    if contato:
        sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('Telefone', contato[0].contato)

    if email:
        sh.cell(row= 2, column= 1).value = (sh.cell(row= 2, column= 1).value).replace('e-mail', email[0].contato)

    sh.cell(row= 2, column= 1).font = styles.Font(name= 'Carlito', bold=True, size= 8)

    # dados aluno
    sh.cell(row= 4, column= 2).value = (sh.cell(row= 4, column= 2).value).replace('turma.etapa', turma.etapa.nome)
    sh.cell(row= 4, column= 2).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 5, column= 2).value = (sh.cell(row= 5, column= 2).value).replace('aluno.nome', aluno.nome)
    sh.cell(row= 5, column= 2).font = styles.Font(name= 'Calibri', size= 9)

    data_nascimento = aluno.nascimento.split('-')
    data_nascimento = f'{data_nascimento[2]}/{data_nascimento[1]}/{data_nascimento[0]}'

    sh.cell(row= 6, column= 1).value = (sh.cell(row= 6, column= 1).value).replace('aluno.data_nascimento', data_nascimento)
    sh.cell(row= 6, column= 1).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 6, column= 4).value = (sh.cell(row= 6, column= 4).value).replace('aluno.naturalidade', str(aluno.naturalidade).replace('None', ''))
    sh.cell(row= 6, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 6, column= 8).value = (sh.cell(row= 6, column= 8).value).replace('aluno.uf', str(aluno.uf).replace('None', ''))
    sh.cell(row= 6, column= 8).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 7, column= 2).value = (sh.cell(row= 7, column= 2).value).replace('aluno.pai', aluno.nome_pai)
    sh.cell(row= 7, column= 2).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 8, column= 2).value = (sh.cell(row= 8, column= 2).value).replace('aluno.mae', aluno.nome_mae)
    sh.cell(row= 8, column= 2).font = styles.Font(name= 'Calibri', size= 9)


    #Relatorio1
    sh.cell(row= 11, column= 4).value = (sh.cell(row= 11, column= 4).value).replace('relatorio1.portugues', relatorio1.portugues)
    sh.cell(row= 11, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 12, column= 4).value = (sh.cell(row= 12, column= 4).value).replace('relatorio1.artes', relatorio1.arte)
    sh.cell(row= 12, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 13, column= 4).value = (sh.cell(row= 13, column= 4).value).replace('relatorio1.ingles', relatorio1.ingles)
    sh.cell(row= 13, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 14, column= 4).value = (sh.cell(row= 14, column= 4).value).replace('relatorio1.ed_fisica', relatorio1.ed_fisica)
    sh.cell(row= 14, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 15, column= 4).value = (sh.cell(row= 13, column= 4).value).replace('relatorio1.matematica', relatorio1.matematica)
    sh.cell(row= 15, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 16, column= 4).value = (sh.cell(row= 16, column= 4).value).replace('relatorio1.fisica', relatorio1.fisica)
    sh.cell(row= 16, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 17, column= 4).value = (sh.cell(row= 17, column= 4).value).replace('relatorio1.quimica', relatorio1.quimica)
    sh.cell(row= 17, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 18, column= 4).value = (sh.cell(row= 18, column= 4).value).replace('relatorio1.biologia', relatorio1.biologia)
    sh.cell(row= 18, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 19, column= 4).value = (sh.cell(row= 19, column= 4).value).replace('relatorio1.historia', relatorio1.historia)
    sh.cell(row= 19, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 20, column= 4).value = (sh.cell(row= 20, column= 4).value).replace('relatorio1.geografia', relatorio1.geografia)
    sh.cell(row= 20, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 21, column= 4).value = (sh.cell(row= 21, column= 4).value).replace('relatorio1.filosofia', relatorio1.filosofia)
    sh.cell(row= 21, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 22, column= 4).value = (sh.cell(row= 22, column= 4).value).replace('relatorio1.sociologia', relatorio1.sociologia)
    sh.cell(row= 22, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 25, column= 4).value = (sh.cell(row= 25, column= 4).value).replace('relatorio1.eletiva', relatorio1.eletiva)
    sh.cell(row= 25, column= 4).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 26, column= 4).value = (sh.cell(row= 26, column= 4).value).replace('relatorio1.pj_vd', relatorio1.projeto_vida)
    sh.cell(row= 26, column= 4).font = styles.Font(name= 'Calibri', size= 9)



    #Relatorio2
    sh.cell(row= 11, column= 6).value = (sh.cell(row= 11, column= 6).value).replace('relatorio2.portugues', relatorio2.portugues)
    sh.cell(row= 11, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 12, column= 6).value = (sh.cell(row= 12, column= 6).value).replace('relatorio2.artes', relatorio2.arte)
    sh.cell(row= 12, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 13, column= 6).value = (sh.cell(row= 13, column= 6).value).replace('relatorio2.ingles', relatorio2.ingles)
    sh.cell(row= 13, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 14, column= 6).value = (sh.cell(row= 14, column= 6).value).replace('relatorio2.ed_fisica', relatorio2.ed_fisica)
    sh.cell(row= 14, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 15, column= 6).value = (sh.cell(row= 15, column= 6).value).replace('relatorio2.matematica', relatorio2.matematica)
    sh.cell(row= 15, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 16, column= 6).value = (sh.cell(row= 16, column= 6).value).replace('relatorio2.fisica', relatorio2.fisica)
    sh.cell(row= 16, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 17, column= 6).value = (sh.cell(row= 17, column= 6).value).replace('relatorio2.quimica', relatorio2.quimica)
    sh.cell(row= 17, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 18, column= 6).value = (sh.cell(row= 18, column= 6).value).replace('relatorio2.biologia', relatorio2.biologia)
    sh.cell(row= 18, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 19, column= 6).value = (sh.cell(row= 19, column= 6).value).replace('relatorio2.historia', relatorio2.historia)
    sh.cell(row= 19, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 20, column= 6).value = (sh.cell(row= 20, column= 6).value).replace('relatorio2.geografia', relatorio2.geografia)
    sh.cell(row= 20, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 21, column= 6).value = (sh.cell(row= 21, column= 6).value).replace('relatorio2.filosofia', relatorio2.filosofia)
    sh.cell(row= 21, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 22, column= 6).value = (sh.cell(row= 22, column= 6).value).replace('relatorio2.sociologia', relatorio2.sociologia)
    sh.cell(row= 22, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 24, column= 6).value = (sh.cell(row= 24, column= 6).value).replace('relatorio2.espanhol', relatorio2.espanhol)
    sh.cell(row= 24, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 25, column= 6).value = (sh.cell(row= 25, column= 6).value).replace('relatorio2.eletiva', relatorio2.eletiva)
    sh.cell(row= 25, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 26, column= 6).value = (sh.cell(row= 26, column= 6).value).replace('relatorio2.pj_vd', relatorio2.projeto_vida)
    sh.cell(row= 26, column= 6).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 28, column= 6).value = (sh.cell(row= 28, column= 6).value).replace('relatorio2.ling_tec', relatorio2.linguagens_tecnologias)
    sh.cell(row= 28, column= 6).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio2.linguagens_tecnologias != '-':
        sh.cell(row= 28, column= 7).value = '200'
        sh.cell(row= 28, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 28, column= 7).value = '-'
        sh.cell(row= 28, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 29, column= 6).value = (sh.cell(row= 29, column= 6).value).replace('relatorio2.mat_tec', relatorio2.matematica_tecnologias)
    sh.cell(row= 29, column= 6).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio2.matematica_tecnologias != '-':
        sh.cell(row= 29, column= 7).value = '200'
        sh.cell(row= 29, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 29, column= 7).value = '-'
        sh.cell(row= 29, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 30, column= 6).value = (sh.cell(row= 30, column= 6).value).replace('relatorio2.cien_nat', relatorio2.ciencias_natureza)
    sh.cell(row= 30, column= 6).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio2.ciencias_natureza != '-':
        sh.cell(row= 30, column= 7).value = '200'
        sh.cell(row= 30, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 30, column= 7).value = '-'
        sh.cell(row= 30, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 31, column= 6).value = (sh.cell(row= 31, column= 6).value).replace('relatorio2.cien_hum', relatorio2.ciencias_humanas)
    sh.cell(row= 31, column= 6).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio2.ciencias_humanas != '-':
        sh.cell(row= 31, column= 7).value = '200'
        sh.cell(row= 31, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 31, column= 7).value = '-'
        sh.cell(row= 31, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 32, column= 6).value = (sh.cell(row= 32, column= 6).value).replace('relatorio2.form_prof', relatorio2.formacao_tecnica)
    sh.cell(row= 32, column= 6).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio2.formacao_tecnica != '-':
        sh.cell(row= 32, column= 7).value = '200'
        sh.cell(row= 32, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 32, column= 7).value = '-'
        sh.cell(row= 32, column= 7).alignment = styles.Alignment(horizontal='center', vertical= 'center')


    # Relatorio3
    sh.cell(row= 11, column= 8).value = (sh.cell(row= 11, column= 8).value).replace('relatorio3.portugues', relatorio3.portugues)
    sh.cell(row= 11, column= 8).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 14, column= 8).value = (sh.cell(row= 14, column= 8).value).replace('relatorio3.ed_fisica', relatorio3.ed_fisica)
    sh.cell(row= 14, column= 8).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 15, column= 8).value = (sh.cell(row= 15, column= 8).value).replace('relatorio3.matematica', relatorio3.matematica)
    sh.cell(row= 15, column= 8).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 27, column= 8).value = (sh.cell(row= 27, column= 8).value).replace('relatorio3.pos_medio', relatorio3.pos_medio)
    sh.cell(row= 27, column= 8).font = styles.Font(name= 'Calibri', size= 9)

    sh.cell(row= 28, column= 8).value = (sh.cell(row= 28, column= 8).value).replace('relatorio3.ling_tec', relatorio3.linguagens_tecnologias)
    sh.cell(row= 28, column= 8).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio3.linguagens_tecnologias != '-':
        sh.cell(row= 28, column= 9).value = '600'
        sh.cell(row= 28, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 28, column= 9).value = '-'
        sh.cell(row= 28, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 29, column= 8).value = (sh.cell(row= 29, column= 8).value).replace('relatorio3.mat_tec', relatorio3.matematica_tecnologias)
    sh.cell(row= 29, column= 8).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio3.matematica_tecnologias != '-':
        sh.cell(row= 29, column= 9).value = '600'
        sh.cell(row= 29, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 29, column= 9).value = '-'
        sh.cell(row= 29, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 30, column= 8).value = (sh.cell(row= 30, column= 8).value).replace('relatorio3.cien_nat', relatorio3.ciencias_natureza)
    sh.cell(row= 30, column= 8).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio3.ciencias_natureza != '-':
        sh.cell(row= 30, column= 9).value = '600'
        sh.cell(row= 30, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 30, column= 9).value = '-'
        sh.cell(row= 30, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 31, column= 8).value = (sh.cell(row= 31, column= 8).value).replace('relatorio3.cien_hum', relatorio3.ciencias_humanas)
    sh.cell(row= 31, column= 8).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio3.ciencias_humanas != '-':
        sh.cell(row= 31, column= 9).value = '600'
        sh.cell(row= 31, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 31, column= 9).value = '-'
        sh.cell(row= 31, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    sh.cell(row= 32, column= 8).value = (sh.cell(row= 32, column= 8).value).replace('relatorio3.form_prof', relatorio3.formacao_tecnica)
    sh.cell(row= 32, column= 8).font = styles.Font(name= 'Calibri', size= 9)
    if relatorio3.formacao_tecnica != '-':
        sh.cell(row= 32, column= 9).value = '600'
        sh.cell(row= 32, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')
    else:
        sh.cell(row= 32, column= 9).value = '-'
        sh.cell(row= 32, column= 9).alignment = styles.Alignment(horizontal='center', vertical= 'center')

    c = 0
    for i in range(40, 43, 1):
        sh.cell(row= i, column= 5).value = (sh.cell(row= i, column= 5).value).replace('NOME DA ESCOLA', escola.escola.nome_escola)
        sh.cell(row= i, column= 5).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        sh.cell(row= i, column= 5).font = styles.Font(name= 'Carlito', size= 5)

        sh.cell(row= i, column= 5).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= i, column= 10).value = (sh.cell(row= i, column= 10).value).replace(f'relatorio{c+1}.resultado', resultados[c])
        sh.cell(row= i, column= 10).alignment = styles.Alignment(horizontal='center', vertical= 'center')

        sh.cell(row= i, column= 10).value = (sh.cell(row= i, column= 10).value).replace(f'relatorio{c+1}.resultado', resultados[c])
        sh.cell(row= i, column= 10).alignment = styles.Alignment(horizontal='center', vertical= 'center')
        c += 1

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Aluno Parcial.xlsx"'

    # Salva os dados na planilha.
    wb.save(response)

    return response